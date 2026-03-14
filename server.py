"""
Flask app: UI and API for decision maker finder.
Run from repo root: python server.py
"""
import os
from pathlib import Path

# Load .env from repo root so HUNTER_API_KEY is available
from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent / ".env")
import csv
import io
import json
from flask import Flask, render_template, request, jsonify, Response, stream_with_context
from openpyxl import load_workbook

from backend import db
from backend.hunter import _get_api_key, company_to_domain, find_email
from backend.pipeline import run_pipeline, run_pipeline_events

# Default job titles
DEFAULT_TITLES = [
    "Head of HR",
    "Chief People Officer",
    "VP People",
    "Head of People",
    "Head of Engineering",
    "VP Engineering",
    "CTO",
    "Head of Learning & Development",
    "Director of People & Culture",
    "Head of Employee Experience",
]

app = Flask(__name__, template_folder="templates", static_folder="static")


def _cors_headers():
    origin = request.environ.get("HTTP_ORIGIN", "")
    allowed_patterns = [
        "localhost",
        "127.0.0.1",
        ".railway.app",
        ".up.railway.app",
        "trustle.online",
    ]
    if origin and any(pattern in origin for pattern in allowed_patterns):
        return {
            "Access-Control-Allow-Origin": origin,
            "Access-Control-Allow-Methods": "GET, POST, PATCH, DELETE, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type",
        }
    return {}


def _parse_bool_filter(val):
    """Return True/False for '1'/'0' style query params, None for all."""
    if not val or val == "":
        return None
    if str(val).lower() in ("1", "true", "yes"):
        return True
    if str(val).lower() == "0":
        return False
    return None


@app.route("/api/auth/login", methods=["POST", "OPTIONS"])
def api_auth_login():
    """Authenticate user with username and password."""
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers.update(_cors_headers())
        return resp
    
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    
    expected_username = os.environ.get("PORTAL_USERNAME", "admin")
    expected_password = os.environ.get("PORTAL_PASSWORD", "trustle2024")
    
    if username == expected_username and password == expected_password:
        resp = jsonify({"success": True, "message": "Login successful"})
        resp.headers.update(_cors_headers())
        return resp
    else:
        resp = jsonify({"success": False, "error": "Invalid username or password"})
        resp.headers.update(_cors_headers())
        return resp, 401


@app.route("/api/auth/verify", methods=["GET", "OPTIONS"])
def api_auth_verify():
    """Verify if the current session is valid (for client-side token validation)."""
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers.update(_cors_headers())
        return resp
    
    resp = jsonify({"valid": True})
    resp.headers.update(_cors_headers())
    return resp


@app.before_request
def cors_preflight():
    if request.method == "OPTIONS":
        return Response(status=204, headers=_cors_headers())


@app.after_request
def add_cors(resp):
    for k, v in _cors_headers().items():
        resp.headers[k] = v
    return resp


@app.route("/")
def index():
    return render_template("index.html", default_titles="\n".join(DEFAULT_TITLES))


@app.route("/api/run", methods=["POST"])
def api_run():
    data = request.get_json() or {}
    titles_raw = data.get("job_titles") or data.get("jobTitles") or ""
    titles = [t.strip() for t in titles_raw.splitlines() if t.strip()]
    if not titles:
        titles = DEFAULT_TITLES
    hunter_enabled = data.get("hunter_enabled", True)
    result = run_pipeline(titles, hunter_enabled=hunter_enabled)
    return jsonify(result)


def _stream_run_events():
    data = request.get_json() or {}
    titles_raw = data.get("job_titles") or data.get("jobTitles") or ""
    titles = [t.strip() for t in titles_raw.splitlines() if t.strip()]
    if not titles:
        titles = DEFAULT_TITLES
    hunter_enabled = data.get("hunter_enabled", True)
    linkedin_only = data.get("linkedin_only", False)
    try:
        for event in run_pipeline_events(titles, hunter_enabled=hunter_enabled, linkedin_only=linkedin_only):
            yield "data: {}\n\n".format(json.dumps(event))
    except Exception as e:
        yield "data: {}\n\n".format(json.dumps({
            "type": "done",
            "saved": 0,
            "with_email": 0,
            "errors": [str(e)],
            "run_id": None,
        }))


@app.route("/api/run/stream", methods=["POST"])
def api_run_stream():
    return Response(
        stream_with_context(_stream_run_events()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/config")
def api_config():
    return jsonify({"hunter_configured": bool(_get_api_key())})


@app.route("/api/runs")
def api_runs():
    limit = min(int(request.args.get("limit", 50)), 100)
    rows = db.list_runs(limit=limit)
    return jsonify({"runs": rows})


@app.route("/api/contacts", methods=["GET", "POST"])
def api_contacts():
    if request.method == "POST":
        data = request.get_json() or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name is required"}), 400
        email = (data.get("email") or "").strip() or None
        job_title = (data.get("job_title") or "").strip() or None
        company = (data.get("company") or "").strip() or None
        run_id = (data.get("run_id") or "").strip() or None
        source_url = (data.get("source_url") or "").strip() or None
        try:
            contact_id = db.insert_contact(name=name, email=email, job_title=job_title, company=company, run_id=run_id, source_url=source_url)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        row = db.get_contact(contact_id)
        return jsonify({"contact": row}), 201

    job_title = request.args.get("job_title", "").strip() or None
    company = request.args.get("company", "").strip() or None
    has_email_arg = request.args.get("has_email", "").lower()
    has_email_only = has_email_arg in ("1", "true", "yes")
    no_email_only = has_email_arg == "0"
    run_id = request.args.get("run_id", "").strip() or None
    name = request.args.get("name", "").strip() or None
    validated = _parse_bool_filter(request.args.get("validated", ""))
    contacted = _parse_bool_filter(request.args.get("contacted", ""))
    responded = _parse_bool_filter(request.args.get("responded", ""))
    unemployed_arg = request.args.get("unemployed", "").lower()
    unemployed_filter = True if unemployed_arg in ("1", "true", "yes") else (False if unemployed_arg == "0" else None)
    has_source_arg = request.args.get("has_source", "").lower()
    has_source_only = has_source_arg in ("1", "true", "yes")
    no_source_only = has_source_arg == "0"
    limit = min(int(request.args.get("limit", 500)), 2000)
    rows = db.list_contacts(
        job_title_filter=job_title,
        company_filter=company,
        has_email_only=has_email_only,
        no_email_only=no_email_only,
        run_id=run_id,
        name_filter=name,
        validated=validated,
        contacted=contacted,
        responded=responded,
        unemployed_filter=unemployed_filter,
        has_source_only=has_source_only,
        no_source_only=no_source_only,
        limit=limit,
    )
    return jsonify({"contacts": rows})


@app.route("/api/contacts/stats")
def api_contacts_stats():
    """Return aggregate stats for contacts with same query params as GET /api/contacts."""
    job_title = request.args.get("job_title", "").strip() or None
    company = request.args.get("company", "").strip() or None
    has_email_arg = request.args.get("has_email", "").lower()
    has_email_only = has_email_arg in ("1", "true", "yes")
    no_email_only = has_email_arg == "0"
    run_id = request.args.get("run_id", "").strip() or None
    name = request.args.get("name", "").strip() or None
    validated = _parse_bool_filter(request.args.get("validated", ""))
    contacted = _parse_bool_filter(request.args.get("contacted", ""))
    responded = _parse_bool_filter(request.args.get("responded", ""))
    unemployed_arg = request.args.get("unemployed", "").lower()
    unemployed_filter = True if unemployed_arg in ("1", "true", "yes") else (False if unemployed_arg == "0" else None)
    has_source_arg = request.args.get("has_source", "").lower()
    has_source_only = has_source_arg in ("1", "true", "yes")
    no_source_only = has_source_arg == "0"
    stats = db.contact_stats(
        job_title_filter=job_title,
        company_filter=company,
        has_email_only=has_email_only,
        no_email_only=no_email_only,
        run_id=run_id,
        name_filter=name,
        validated=validated,
        contacted=contacted,
        responded=responded,
        unemployed_filter=unemployed_filter,
        has_source_only=has_source_only,
        no_source_only=no_source_only,
    )
    return jsonify(stats)


def _first_last(name):
    """Split full name into first and last for Hunter."""
    parts = (name or "").strip().split(maxsplit=1)
    if len(parts) >= 2:
        return parts[0], parts[1]
    if len(parts) == 1:
        return parts[0], ""
    return "", ""


@app.route("/api/contacts/enrich", methods=["POST"])
def api_contacts_enrich():
    """Bulk enrich contacts without email using Hunter.io. Body: { contact_ids: [1,2,3] } or { filter: "no_email" }."""
    data = request.get_json() or {}
    contact_ids = data.get("contact_ids")
    if data.get("filter") == "no_email":
        rows = db.list_contacts(no_email_only=True, limit=500)
        contact_ids = [r["id"] for r in rows]
    if not contact_ids:
        return jsonify({"enriched": 0, "skipped": 0, "errors": []})
    enriched = 0
    skipped = 0
    errors = []
    for cid in contact_ids:
        try:
            c = db.get_contact(cid)
            if not c:
                errors.append({"id": cid, "message": "Contact not found"})
                continue
            email_val = (c.get("email") or "").strip()
            if email_val and "@" in email_val:
                skipped += 1
                continue
            company = (c.get("company") or "").strip()
            name = (c.get("name") or "").strip()
            if not company or not name:
                skipped += 1
                continue
            domain, domain_err = company_to_domain(company)
            if domain_err:
                errors.append({"id": cid, "message": domain_err})
                continue
            if not domain:
                skipped += 1
                continue
            first_name, last_name = _first_last(name)
            email, email_err = find_email(domain, first_name, last_name)
            if email_err:
                errors.append({"id": cid, "message": email_err})
                continue
            if email and "@" in email:
                db.update_contact(cid, email=email)
                enriched += 1
            else:
                skipped += 1
        except Exception as e:
            errors.append({"id": cid, "message": str(e)})
    return jsonify({"enriched": enriched, "skipped": skipped, "errors": errors})


@app.route("/api/contacts/<int:contact_id>", methods=["PATCH", "DELETE"])
def api_contact_by_id(contact_id):
    if request.method == "DELETE":
        deleted = db.delete_contact(contact_id)
        if not deleted:
            return jsonify({"error": "Not found"}), 404
        return jsonify({"deleted": True})
    if request.method == "PATCH":
        data = request.get_json() or {}
        name = data.get("name")
        email = data.get("email")
        job_title = data.get("job_title")
        company = data.get("company")
        validated = data.get("validated")
        contacted = data.get("contacted")
        responded = data.get("responded")
        unemployed = data.get("unemployed")
        alternative_email = data.get("alternative_email")
        source_url = data.get("source_url")
        if name is not None:
            name = str(name).strip()
        if email is not None:
            email = str(email).strip()
        if job_title is not None:
            job_title = str(job_title).strip()
        if company is not None:
            company = str(company).strip()
        if alternative_email is not None:
            alternative_email = str(alternative_email).strip()
        if source_url is not None:
            source_url = str(source_url).strip()
        if validated is not None:
            validated = bool(validated)
        if contacted is not None:
            contacted = bool(contacted)
        if responded is not None:
            responded = bool(responded)
        if unemployed is not None:
            unemployed = bool(unemployed)
        updated = db.update_contact(
            contact_id,
            name=name,
            email=email,
            job_title=job_title,
            company=company,
            validated=validated,
            contacted=contacted,
            responded=responded,
            unemployed=unemployed,
            alternative_email=alternative_email,
            source_url=source_url,
        )
        if not updated:
            return jsonify({"error": "Not found"}), 404
        return jsonify({"updated": True, "email": email, "job_title": job_title, "company": company})


# Column mapping for flexible xlsx import
COLUMN_MAPPINGS = {
    "name": ["name", "full name", "fullname", "contact name", "person", "contact"],
    "email": ["email", "email address", "e-mail", "work email", "primary email", "mail"],
    "job_title": ["job title", "title", "position", "role", "job"],
    "company": ["company", "company name", "organization", "employer", "firm"],
    "company_domain": ["domain", "company domain", "website"],
    "source_url": ["linkedin", "linkedin url", "source url", "source", "profile url", "url", "link"],
    "validated": ["validated", "valid", "verified"],
    "contacted": ["contacted"],
    "responded": ["responded"],
    "unemployed": ["unemployed"],
    "alternative_email": ["alternative email", "alt email", "secondary email", "alternate email", "other email", "alt. email"],
}


def _parse_bool_cell(val):
    """Return True if value is Y/yes/1/true, else False."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("Y", "YES", "1", "TRUE")


def _normalize_header(h):
    """Normalize header for matching: lowercase, strip, remove underscores."""
    if not h:
        return ""
    return str(h).strip().lower().replace("_", " ")


def _detect_column_mappings(header_row):
    """
    Detect which columns map to which DB fields.
    Returns dict: {db_field: (col_index, original_header_name)}
    Uses a two-pass approach: exact matches first, then partial matches.
    """
    if not header_row:
        return {}
    
    col_map = {}
    used_cols = set()
    
    # First pass: exact matches only
    for col_idx, header in enumerate(header_row):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        
        for db_field, variations in COLUMN_MAPPINGS.items():
            if db_field in col_map:
                continue
            for variation in variations:
                if normalized == variation:
                    col_map[db_field] = (col_idx, str(header).strip())
                    used_cols.add(col_idx)
                    break
    
    # Second pass: partial matches for remaining columns
    # Sort variations by length (longest first) to prefer more specific matches
    for col_idx, header in enumerate(header_row):
        if col_idx in used_cols:
            continue
        normalized = _normalize_header(header)
        if not normalized:
            continue
        
        best_match = None
        best_match_len = 0
        
        for db_field, variations in COLUMN_MAPPINGS.items():
            if db_field in col_map:
                continue
            for variation in variations:
                if variation in normalized and len(variation) > best_match_len:
                    best_match = db_field
                    best_match_len = len(variation)
        
        if best_match:
            col_map[best_match] = (col_idx, str(header).strip())
            used_cols.add(col_idx)
    
    return col_map


def _get_cell_value(row, col_idx):
    """Safely get a cell value from a row."""
    if col_idx is None or col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None:
        return None
    return str(val).strip() if val else ""


def _normalize_source_url(url):
    """Ensure URL has a scheme; linkedin.com/in/... -> https://www.linkedin.com/in/..."""
    if not url or not isinstance(url, str):
        return ""
    s = url.strip()
    if not s:
        return ""
    if s.startswith("http://") or s.startswith("https://"):
        return s
    if s.startswith("linkedin.com/"):
        return "https://www." + s
    if s.startswith("www."):
        return "https://" + s
    return "https://" + s


def _find_existing_contact(email, company, name):
    """
    Find existing contact: first by email, then by (company, name).
    Returns contact id or None.
    """
    if email and "@" in email:
        path = db.get_db_path()
        import sqlite3
        conn = sqlite3.connect(path)
        cur = conn.execute(
            "SELECT id FROM contacts WHERE email IS NOT NULL AND TRIM(email) != '' AND LOWER(TRIM(email)) = LOWER(?)",
            (email.strip(),),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            return row[0]
    
    if company and name:
        return db.find_contact_id_by_company_and_name(company, name)
    
    return None


def _process_import_row(row, row_idx, col_map):
    """
    Process a single data row with flexible column mapping.
    Returns (created_delta, updated_delta, skipped_delta, error_dict or None, affected_ids).
    """
    try:
        vals = list(row) if row else []
        if not vals or all(v is None or str(v).strip() == "" for v in vals):
            return (0, 0, 1, None, [])
        
        name_info = col_map.get("name")
        name = _get_cell_value(vals, name_info[0]) if name_info else None
        
        email_info = col_map.get("email")
        email = _get_cell_value(vals, email_info[0]) if email_info else None
        
        company_info = col_map.get("company")
        company = _get_cell_value(vals, company_info[0]) if company_info else None
        
        if not name and not (company and email):
            return (0, 0, 1, None, [])
        
        job_title_info = col_map.get("job_title")
        job_title = _get_cell_value(vals, job_title_info[0]) if job_title_info else None
        
        company_domain_info = col_map.get("company_domain")
        company_domain = _get_cell_value(vals, company_domain_info[0]) if company_domain_info else None
        
        source_url_info = col_map.get("source_url")
        source_url_raw = _get_cell_value(vals, source_url_info[0]) if source_url_info else None
        source_url = _normalize_source_url(source_url_raw) if source_url_raw else None
        
        validated_info = col_map.get("validated")
        validated = _parse_bool_cell(vals[validated_info[0]] if validated_info and validated_info[0] < len(vals) else None)
        
        contacted_info = col_map.get("contacted")
        contacted = _parse_bool_cell(vals[contacted_info[0]] if contacted_info and contacted_info[0] < len(vals) else None)
        
        responded_info = col_map.get("responded")
        responded = _parse_bool_cell(vals[responded_info[0]] if responded_info and responded_info[0] < len(vals) else None)
        
        unemployed_info = col_map.get("unemployed")
        unemployed = _parse_bool_cell(vals[unemployed_info[0]] if unemployed_info and unemployed_info[0] < len(vals) else None)
        
        alt_email_info = col_map.get("alternative_email")
        alternative_email = _get_cell_value(vals, alt_email_info[0]) if alt_email_info else None
        
        existing_id = _find_existing_contact(email, company, name)
        
        if existing_id:
            update_kwargs = {}
            if name:
                update_kwargs["name"] = name
            if email:
                update_kwargs["email"] = email
            if job_title:
                update_kwargs["job_title"] = job_title
            if company:
                update_kwargs["company"] = company
            if source_url:
                update_kwargs["source_url"] = source_url
            if validated_info:
                update_kwargs["validated"] = validated
            if contacted_info:
                update_kwargs["contacted"] = contacted
            if responded_info:
                update_kwargs["responded"] = responded
            if unemployed_info:
                update_kwargs["unemployed"] = unemployed
            if alternative_email:
                update_kwargs["alternative_email"] = alternative_email
            
            if update_kwargs:
                db.update_contact(existing_id, **update_kwargs)
            return (0, 1, 0, None, [existing_id])
        else:
            if not name:
                return (0, 0, 1, None, [])
            new_id = db.insert_contact(
                name=name,
                email=email or None,
                job_title=job_title or None,
                company=company or None,
                validated=validated,
                source_url=source_url or None,
            )
            if contacted_info or responded_info or unemployed_info or alternative_email:
                update_kwargs = {}
                if contacted_info:
                    update_kwargs["contacted"] = contacted
                if responded_info:
                    update_kwargs["responded"] = responded
                if unemployed_info:
                    update_kwargs["unemployed"] = unemployed
                if alternative_email:
                    update_kwargs["alternative_email"] = alternative_email
                if update_kwargs:
                    db.update_contact(new_id, **update_kwargs)
            return (1, 0, 0, None, [new_id])
    except Exception as e:
        return (0, 0, 0, {"row": row_idx, "message": str(e)}, [])


def _import_xlsx_to_contacts(file_stream):
    """
    Parse xlsx with flexible column detection and upsert contacts.
    Returns (created, updated, skipped, errors, contacts, column_mappings).
    """
    wb = load_workbook(filename=io.BytesIO(file_stream.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return 0, 0, 0, [{"row": 0, "message": "Sheet is empty"}], [], {}
    
    header_row = rows[0]
    col_map = _detect_column_mappings(header_row)
    
    if not col_map.get("name") and not (col_map.get("company") and col_map.get("email")):
        return 0, 0, 0, [{"row": 0, "message": "Could not detect required columns. Need 'Name' or both 'Company' and 'Email'."}], [], {}
    
    column_mappings = {info[1]: field for field, info in col_map.items()}
    
    data_rows = rows[1:]
    created = 0
    updated = 0
    skipped = 0
    errors = []
    affected_ids = []
    
    for row_idx, row in enumerate(data_rows, start=2):
        c_d, u_d, s_d, err, ids = _process_import_row(row, row_idx, col_map)
        created += c_d
        updated += u_d
        skipped += s_d
        if err:
            errors.append(err)
        affected_ids.extend(ids)
    
    contacts = db.get_contacts_by_ids(affected_ids) if affected_ids else []
    return created, updated, skipped, errors, contacts, column_mappings


def _stream_import_xlsx(file_stream):
    """Yield NDJSON lines: progress events then a done event with contacts and column_mappings."""
    wb = load_workbook(filename=io.BytesIO(file_stream.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        yield json.dumps({"type": "done", "created": 0, "updated": 0, "skipped": 0, "errors": [{"row": 0, "message": "Sheet is empty"}], "contacts": [], "column_mappings": {}}) + "\n"
        return
    
    header_row = rows[0]
    col_map = _detect_column_mappings(header_row)
    
    if not col_map.get("name") and not (col_map.get("company") and col_map.get("email")):
        yield json.dumps({"type": "done", "created": 0, "updated": 0, "skipped": 0, "errors": [{"row": 0, "message": "Could not detect required columns. Need 'Name' or both 'Company' and 'Email'."}], "contacts": [], "column_mappings": {}}) + "\n"
        return
    
    column_mappings = {info[1]: field for field, info in col_map.items()}
    
    data_rows = rows[1:]
    total = len(data_rows)
    created, updated, skipped = 0, 0, 0
    errors = []
    affected_ids = []
    
    for i, row in enumerate(data_rows):
        row_idx = i + 2
        c_d, u_d, s_d, err, ids = _process_import_row(row, row_idx, col_map)
        created += c_d
        updated += u_d
        skipped += s_d
        if err:
            errors.append(err)
        affected_ids.extend(ids)
        yield json.dumps({"type": "progress", "processed": i + 1, "total": total}) + "\n"
    
    contacts = db.get_contacts_by_ids(affected_ids) if affected_ids else []
    yield json.dumps({"type": "done", "created": created, "updated": updated, "skipped": skipped, "errors": errors, "contacts": contacts, "column_mappings": column_mappings}) + "\n"


@app.route("/api/imports/upload", methods=["POST"])
def api_imports_upload():
    """Accept .xlsx file with flexible column detection. If stream=1, return NDJSON stream with progress; else return JSON result."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded", "created": 0, "updated": 0, "skipped": 0, "errors": [], "column_mappings": {}}), 400
    f = request.files["file"]
    if not f.filename or not (f.filename.endswith(".xlsx") or f.filename.endswith(".xls")):
        return jsonify({"error": "Please upload an .xlsx file", "created": 0, "updated": 0, "skipped": 0, "errors": [], "column_mappings": {}}), 400
    stream = request.args.get("stream", "").lower() in ("1", "true", "yes")
    if stream:
        try:
            return Response(
                stream_with_context(_stream_import_xlsx(f)),
                mimetype="application/x-ndjson",
                headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
            )
        except Exception as e:
            return jsonify({"error": str(e), "created": 0, "updated": 0, "skipped": 0, "errors": [], "column_mappings": {}}), 500
    try:
        created, updated, skipped, errors, contacts, column_mappings = _import_xlsx_to_contacts(f)
        return jsonify({"created": created, "updated": updated, "skipped": skipped, "errors": errors, "contacts": contacts, "column_mappings": column_mappings})
    except Exception as e:
        return jsonify({"error": str(e), "created": 0, "updated": 0, "skipped": 0, "errors": [], "contacts": [], "column_mappings": {}}), 500


@app.route("/api/export")
def api_export():
    job_title = request.args.get("job_title", "").strip() or None
    company = request.args.get("company", "").strip() or None
    has_email_arg = request.args.get("has_email", "").lower()
    has_email_only = has_email_arg in ("1", "true", "yes")
    no_email_only = has_email_arg == "0"
    run_id = request.args.get("run_id", "").strip() or None
    name = request.args.get("name", "").strip() or None
    validated = _parse_bool_filter(request.args.get("validated", ""))
    contacted = _parse_bool_filter(request.args.get("contacted", ""))
    responded = _parse_bool_filter(request.args.get("responded", ""))
    unemployed_arg = request.args.get("unemployed", "").lower()
    unemployed_filter = True if unemployed_arg in ("1", "true", "yes") else (False if unemployed_arg == "0" else None)
    has_source_arg = request.args.get("has_source", "").lower()
    has_source_only = has_source_arg in ("1", "true", "yes")
    no_source_only = has_source_arg == "0"
    rows = db.list_contacts(
        job_title_filter=job_title,
        company_filter=company,
        has_email_only=has_email_only,
        no_email_only=no_email_only,
        run_id=run_id,
        name_filter=name,
        validated=validated,
        contacted=contacted,
        responded=responded,
        unemployed_filter=unemployed_filter,
        has_source_only=has_source_only,
        no_source_only=no_source_only,
        limit=10000,
    )
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["name", "email", "job_title", "company", "company_domain", "source_url", "validated", "contacted", "responded", "created_at"])
    for r in rows:
        w.writerow([
            r.get("name"),
            r.get("email"),
            r.get("job_title"),
            r.get("company"),
            r.get("company_domain"),
            r.get("source_url"),
            "yes" if r.get("validated") else "no",
            "yes" if r.get("contacted") else "no",
            "yes" if r.get("responded") else "no",
            r.get("created_at"),
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=contacts.csv"},
    )


if __name__ == "__main__":
    db.init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
